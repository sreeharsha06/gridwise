from __future__ import annotations
from typing import List, Tuple, Optional
from openpyxl import load_workbook
from gridwise.core.model import Sheet, Cell
from gridwise.core.utils import idx_to_addr, infer_dtype

def from_xlsx_rich(path: str, sheet_name: str | None = None) -> Sheet:
    wb = load_workbook(filename=path, data_only=True, read_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    nrows = max_row
    ncols = max_col

    merged_regions: List[Tuple[int, int, int, int]] = []
    for mr in ws.merged_cells.ranges:
        r1, c1, r2, c2 = mr.min_row - 1, mr.min_col - 1, mr.max_row - 1, mr.max_col - 1
        merged_regions.append((r1, c1, r2, c2))

    frozen_rows = 0
    frozen_cols = 0
    if ws.freeze_panes:
        fr = ws.freeze_panes
        frozen_rows = (fr.row or 1) - 1 if fr.row else 0
        frozen_cols = (fr.col_idx or 1) - 1 if getattr(fr, "col_idx", None) else 0

    cells: List[Cell] = []
    header_row_index: Optional[int] = None
    for i in range(nrows):
        if any((ws.cell(row=i + 1, column=j + 1).value not in (None, "")) for j in range(ncols)):
            header_row_index = i
            break

    for i in range(nrows):
        for j in range(ncols):
            xl = ws.cell(row=i + 1, column=j + 1)
            val = xl.value
            nfs = xl.number_format if xl.number_format else None
            addr = idx_to_addr(i, j)
            fmt = "header" if (header_row_index is not None and i == header_row_index) else nfs
            cells.append(
                Cell(row=i, col=j, address=addr, value=val, dtype=infer_dtype(val), fmt=fmt)
            )

    return Sheet(
        name=ws.title,
        nrows=nrows,
        ncols=ncols,
        cells=cells,
        merged_regions=merged_regions or None,
        frozen=(frozen_rows, frozen_cols) if (frozen_rows or frozen_cols) else None,
    )
